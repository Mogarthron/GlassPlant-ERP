import datetime
from datetime import date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color
import calendar
import pandas as pd
from pandas import DataFrame, Series

import os.path

if os.path.isfile('Karta1.xlsx'):
    os.remove('Karta1.xlsx')
if os.path.isfile('Harmonogram_Pracy.xlsx'):
    os.remove('Harmonogram_Pracy.xlsx')


class SheduleModel():

    def __init__(self):

        self.MonthNames = ['Styczeń', 'Luty', 'Marzec', 'Kwiecień', 'Maj', 'Czerwiec',
                           'Lipiec', 'Sierpień', 'Wrzesień', 'Październik', 'Listopad', 'Grudzień']

        self.weekDay = ['poniedziałek', 'wtorek', 'środa',
                        'czwartek', 'piątek', 'sobota', 'niedziela']

        self.r = ['Ranek', 'R']
        self.p = ['Popołudnie', 'P']
        self.n = ['Nocka', 'N']
        self.w = ['Wolne', 'W']
        self.wn = ['Wolna niedziela', 'wn']
        self.dt = ['Dzień wolny wynikający z grafiku', 'dt']
        self.l4 = ['Zwolnienie lekarskie', 'L4']
        self.uw = ['Urlop wypoczynkowy', 'uw']
        self.uo = ['Urlop okolicznościowy', 'uo']

    def SchedulePatern(self, patern):
        '''patern in string like RRRRWPPPPWNNNNWW '''

        _schedulePatern = list()

        for i in patern:
            if(i == 'R'):
                _schedulePatern.append(self.r)
            elif (i == 'P'):
                _schedulePatern.append(self.p)
            elif (i == 'N'):
                _schedulePatern.append(self.n)
            elif (i == 'W'):
                _schedulePatern.append(self.w)

        return _schedulePatern

    def WeekDay(self, data):
        r = list()
        r.append(data)
        r.append(self.weekDay[data.weekday()])
        return r

    def MonthDay(self, year, month):

        data = date(year, month, 1)

        monthDay = list()

        if (data.month == 12):
            while (data.year < year + 1):
                monthDay.append(self.WeekDay(data))
                data = data + datetime.timedelta(days=1)
        else:
            while (data.month < month + 1):
                monthDay.append(self.WeekDay(data))
                data = data + datetime.timedelta(days=1)

        return monthDay


class EmployModel():

    def __init__(self, employ, dateOfSchedule, Sundays, shiftShedule, firstShift):

        self.Employ = employ,
        self.DateOfSchedule = dateOfSchedule
        self.NumberOfSundays = Sundays
        self.ShiftShedule = shiftShedule
        self.FirstShift = firstShift


class WorkSchedule():

    def __init__(self, department, year, month):

        self.__sm = SheduleModel()
        self.Department = department
        self.year = year
        self.month = month

        self.MonthNames = self.__sm.MonthNames
        self.weekDay = self.__sm.weekDay

        self.r = ['Ranek', 'R']
        self.p = ['Popołudnie', 'P']
        self.n = ['Nocka', 'N']
        self.w = ['Wolne', 'W']
        self.wn = ['Wolna niedziela', 'wn']
        self.dt = ['Dzień wolny wynikający z grafiku', 'dt']
        self.l4 = ['Zwolnienie lekarskie', 'L4']
        self.uw = ['Urlop wypoczynkowy', 'uw']
        self.uo = ['Urlop okolicznościowy', 'uo']

        self.__departmentName = 'Wydział: '

    def SchedulePatern(self, patern):
        '''patern in string like RRRRWPPPPWNNNNWW '''

        _schedulePatern = list()

        for i in patern:
            if(i == 'R'):
                _schedulePatern.append(self.r)
            elif (i == 'P'):
                _schedulePatern.append(self.p)
            elif (i == 'N'):
                _schedulePatern.append(self.n)
            elif (i == 'W'):
                _schedulePatern.append(self.w)

        return _schedulePatern

    def WeekDay(self, data):
        r = list()
        r.append(data)
        r.append(self.weekDay[data.weekday()])
        return r

    def MonthDay(self, year, month):

        data = date(year, month, 1)

        monthDay = list()

        if (data.month == 12):
            while (data.year < year + 1):
                monthDay.append(self.WeekDay(data))
                data = data + datetime.timedelta(days=1)
        else:
            while (data.month < month + 1):
                monthDay.append(self.WeekDay(data))
                data = data + datetime.timedelta(days=1)

        return monthDay

    def SetShiftsDiuringMonth(self, schedulePatern, startShift, typ_nazwy=1):
        shift = list()

        for __d in self.MonthDay(self.year, self.month):

            shift.append(schedulePatern[startShift])

            startShift = startShift + 1
            if (len(schedulePatern) - 1 < startShift):
                startShift = 0

        return [x[typ_nazwy] for x in shift]

    def CreateSchedule(self):

        czteroBrygadowka = self.SchedulePatern('RRRRRWW')

        col = ['Sokołowski Andrzej', 'Perkowski Leszek',
               'Janiak Tomasz', 'Szewczul Bogusław']

        __fourshiftShedule = 'RRRRWPPPPWNNNNWW'
        __oneshiftShedule = 'RRRRRWW'

        dateOfShedule = date(self.year, self.month, 1)
        listOfEmployes = [
            EmployModel(col[0], dateOfShedule, 0, __fourshiftShedule, 2),
            EmployModel(col[1], dateOfShedule, 0, __fourshiftShedule, 10),
            EmployModel(col[2], dateOfShedule, 0, __fourshiftShedule, 14),
            EmployModel(col[3], dateOfShedule, 0, __fourshiftShedule, 6)]

        df = DataFrame()

        df['Dzień mc'] = [x[0] for x in self.MonthDay(self.year, self.month)]
        df['Dzień tyg'] = [x[1] for x in self.MonthDay(self.year, self.month)]

        df[col[0]] = self.SetShiftsDiuringMonth(
            czteroBrygadowka, 6)

        df[col[1]] = self.SetShiftsDiuringMonth(
            czteroBrygadowka, 6)

        df[col[2]] = self.SetShiftsDiuringMonth(
            czteroBrygadowka, 6)

        df[col[3]] = self.SetShiftsDiuringMonth(
            czteroBrygadowka, 6)

        print('creating Schedule...')
        print(df.to_string(index=False))
        # df.to_excel('Harmonogram_Pracy.xlsx', index=None)
        print('Schedule printed')


class WorkCard():

    def __init__(self, department, year, month):

        self.Department = department
        self.year = year
        self.month = month

        self.MonthNames = ['Styczeń', 'Luty', 'Marzec', 'Kwiecień', 'Maj', 'Czerwiec',
                           'Lipiec', 'Sierpień', 'Wrzesień', 'Październik', 'Listopad', 'Grudzień']

        self.weekDay = ['poniedziałek', 'wtorek', 'środa',
                        'czwartek', 'piątek', 'sobota', 'niedziela']

        self.r = ['Ranek', 'R']
        self.p = ['Popołudnie', 'P']
        self.n = ['Nocka', 'N']
        self.w = ['Wolne', 'W']
        self.wn = ['Wolna niedziela', 'wn']
        self.dt = ['Dzień wolny wynikający z grafiku', 'dt']
        self.l4 = ['Zwolnienie lekarskie', 'L4']
        self.uw = ['Urlop wypoczynkowy', 'uw']
        self.uo = ['Urlop okolicznościowy', 'uo']

        self.__departmentName = 'Wydział: '

    def SchedulePatern(self, patern):
        '''patern in string like rrrrwppppwnnnnww '''

        _schedulePatern = list()

        for i in patern:
            if(i == 'r'):
                _schedulePatern.append(self.r)
            elif (i == 'p'):
                _schedulePatern.append(self.p)
            elif (i == 'n'):
                _schedulePatern.append(self.n)
            elif (i == 'w'):
                _schedulePatern.append(self.w)

        return _schedulePatern

    def WeekDay(self, data):
        r = list()
        r.append(data)
        r.append(self.weekDay[data.weekday()])
        return r

    def MonthDay(self, year, month):

        data = date(year, month, 1)

        monthDay = list()

        if (data.month == 12):
            while (data.year < year + 1):
                monthDay.append(self.WeekDay(data))
                data = data + datetime.timedelta(days=1)
        else:
            while (data.month < month + 1):
                monthDay.append(self.WeekDay(data))
                data = data + datetime.timedelta(days=1)

        return monthDay

    def SetShiftsDiuringMonth(self, schedulePatern, startShift, typ_nazwy=1):
        shift = list()

        for __d in self.MonthDay(self.year, self.month):

            shift.append(schedulePatern[startShift])

            startShift = startShift + 1
            if (len(schedulePatern) - 1 < startShift):
                startShift = 0

        return shift

    def PrintWorkCardToExcell(self, nameOfEmploy, workSchedule):

        # Department = self.__departmentName + self.Department
        Position = 'Topiarz'
        Salary = '2500 mc'
        TypeOfWork = 'Topienie Szkła'
        SplitingOfOpalGlass = 'Wylewanie opalu'
        startShift = 2

        s = self.SetShiftsDiuringMonth(
            self.SchedulePatern('rrrrwppppwrrrrww'), startShift)

        wb = load_workbook(filename='Karta_pracy.xlsx')
        sheet = wb.active

        sheet['A3'] = self.__departmentName + self.Department
        sheet['G1'] = self.MonthNames[self.month-1]
        sheet['G2'] = self.year
        sheet['M1'] = nameOfEmploy
        sheet['AC1'] = Salary
        sheet['AI1'] = Position

        colNum = 4

        for d in range(calendar.monthrange(self.year, self.month)[1]):

            sheet.cell(row=5, column=colNum, value=d + 1)
            # if (date(year, month, d+1).weekday == 6):
            #     CellFontColor = Font(color='00FFFFFF')
            #     CellFillColor = Color('00808080')
            #     sheet.cell(row=5, column=colNum, value=d + 1).font = CellFontColor
            #     sheet.cell(row=5, column=colNum, value=d + 1).fill = CellFillColor

            if (s[d] == self.r):
                sheet.cell(row=6, column=colNum, value=6)
                sheet.cell(row=7, column=colNum, value=14)
                sheet.cell(row=8, column=colNum, value=8)

            elif (s[d] == self.p):
                sheet.cell(row=6, column=colNum, value=14)
                sheet.cell(row=7, column=colNum, value=22)
                sheet.cell(row=8, column=colNum, value=8)

            elif (s[d] == self.n):
                sheet.cell(row=6, column=colNum, value=22)
                sheet.cell(row=7, column=colNum, value=6)
                sheet.cell(row=8, column=colNum, value=8)
                sheet.cell(row=12, column=colNum, value=8)

            elif (s[d] == self.w):
                sheet.cell(row=6, column=colNum, value='')

            colNum = colNum + 1

        sheet['A8'] = TypeOfWork
        sheet['A9'] = SplitingOfOpalGlass
        sheet['AI8'] = '=Sum(D8:AH8)'
        sheet['AI12'] = '=Sum(D12:AH12)'

        print('Printing...')
        wb.save('Karta1.xlsx')
        print('Work card printed')


##########################################################################################################
##########################################################################################################


ws = WorkSchedule('DT', 2020, 11)
ws.CreateSchedule()
