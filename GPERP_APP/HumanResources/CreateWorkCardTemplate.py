import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

# from openpyxl.utils import units
import os
import string


def patch_worksheet():
    """This monkeypatches Worksheet.merge_cells to remove cell deletion bug
    https://bitbucket.org/openpyxl/openpyxl/issues/365/styling-merged-cells-isnt-working
    Thank you to Sergey Pikhovkin for the fix
    """

    def merge_cells(
        self,
        range_string=None,
        start_row=None,
        start_column=None,
        end_row=None,
        end_column=None,
    ):
        """Set merge on a cell range.  Range is a cell range (e.g. A1:E1)
        This is monkeypatched to remove cell deletion bug
        https://bitbucket.org/openpyxl/openpyxl/issues/365/styling-merged-cells-isnt-working
        """
        if not range_string and not all((start_row, start_column, end_row, end_column)):
            msg = "You have to provide a value either for 'coordinate' or for\
            'start_row', 'start_column', 'end_row' *and* 'end_column'"
            raise ValueError(msg)
        elif not range_string:
            range_string = "%s%s:%s%s" % (
                get_column_letter(start_column),
                start_row,
                get_column_letter(end_column),
                end_row,
            )
        elif ":" not in range_string:
            if COORD_RE.match(range_string):
                return  # Single cell, do nothing
            raise ValueError("Range must be a cell range (e.g. A1:E1)")
        else:
            range_string = range_string.replace("$", "")

        if range_string not in self.merged_cells:
            self.merged_cells.add(range_string)

        # The following is removed by this monkeypatch:

        # min_col, min_row, max_col, max_row = range_boundaries(range_string)
        # rows = range(min_row, max_row+1)
        # cols = range(min_col, max_col+1)
        # cells = product(rows, cols)

        # all but the top-left cell are removed
        # for c in islice(cells, 1, None):
        # if c in self._cells:
        # del self._cells[c]

    # Apply monkey patch
    openpyxl.worksheet.worksheet.Worksheet.merge_cells = merge_cells


patch_worksheet()


class Temp:
    def cm_to_jedn_szer(self, value):
        return value / 0.19

    def cm_to_jedn_wysokosci(self, value):
        return value / 0.035

    def CreateTemp(self):

        wk_name = "./GPERP_APP/HumanResources/Karta_pracy_Template.xlsx"

        if os.path.isfile(wk_name):
            os.remove(wk_name)

        wb = openpyxl.Workbook()
        ws = wb.active

        # print(units.BASE_COL_WIDTH)
        # print(units.DEFAULT_COLUMN_WIDTH)
        # print(units.DEFAULT_ROW_HEIGHT)

        width_1 = 0.54  # cm
        width_2 = 2.83  # cm
        width_3 = 0.57  # cm

        height_1 = 0.94  # 27jedn 0.95cm 45px
        height_2 = 0.34  # 9.6jedn 0.34cm 16px
        height_3 = 0.63  # cm
        height_4 = 0.5
        height_5 = 0.57
        height_6_20 = 0.86
        height_21 = 0.59
        height_22 = 0.54
        height_23 = 0.65
        height_24 = 0.82
        height_25 = 0.48

        # 2.9jedn 0.57cm 27px
        # 3jedn 0.59cm 28px
        # 12jedn 2.39cm 113px
        # 3.5jedn 0.68cm 32px

        ws.column_dimensions["A"].width = self.cm_to_jedn_szer(width_1)
        ws.column_dimensions["B"].width = self.cm_to_jedn_szer(width_1)
        ws.column_dimensions["C"].width = self.cm_to_jedn_szer(width_2)

        for i in range(3, 34):
            if i >= 26:
                c = "A" + string.ascii_uppercase[i - 26]
                ws.column_dimensions[c].width = self.cm_to_jedn_szer(width_3)
            else:
                c = string.ascii_uppercase[i]
                ws.column_dimensions[c].width = self.cm_to_jedn_szer(width_3)

        ws.column_dimensions["AI"].width = self.cm_to_jedn_szer(1.28)
        ws.column_dimensions["AJ"].width = self.cm_to_jedn_szer(1.07)
        ws.column_dimensions["AK"].width = self.cm_to_jedn_szer(1.49)
        ws.column_dimensions["AL"].width = self.cm_to_jedn_szer(3.46)

        ws.row_dimensions["1"].height = self.cm_to_jedn_wysokosci(height_1)
        ws.row_dimensions["2"].height = self.cm_to_jedn_wysokosci(height_2)
        ws.row_dimensions["3"].height = self.cm_to_jedn_wysokosci(height_3)
        ws.row_dimensions["4"].height = self.cm_to_jedn_wysokosci(height_4)
        ws.row_dimensions["5"].height = self.cm_to_jedn_wysokosci(height_5)

        for i in range(6, 21):
            if i > 9 and i <= 19:
                c = "1" + string.digits[i - 10]
                ws.row_dimensions[c].height = self.cm_to_jedn_wysokosci(height_6_20)
            elif i >= 20:
                c = "2" + string.digits[i - 20]
                ws.row_dimensions[c].height = self.cm_to_jedn_wysokosci(height_6_20)
            else:
                c = string.digits[i]
                ws.row_dimensions[c].height = self.cm_to_jedn_wysokosci(height_6_20)

        ws.row_dimensions["21"].height = self.cm_to_jedn_wysokosci(height_21)
        ws.row_dimensions["22"].height = self.cm_to_jedn_wysokosci(height_22)
        ws.row_dimensions["23"].height = self.cm_to_jedn_wysokosci(height_23)
        ws.row_dimensions["24"].height = self.cm_to_jedn_wysokosci(height_24)
        ws.row_dimensions["25"].height = self.cm_to_jedn_wysokosci(height_25)

        # border = Border(
        #     left=Side(border_style=None, color="FF000000"),
        #     right=Side(border_style=None, color="FF000000"),
        #     top=Side(border_style=None, color="FF000000"),
        #     bottom=Side(border_style=None, color="FF000000"),
        #     diagonal=Side(border_style=None, color="FF000000"),
        #     diagonal_direction=0,
        #     outline=Side(border_style=None, color="FF000000"),
        #     vertical=Side(border_style=None, color="FF000000"),
        #     horizontal=Side(border_style=None, color="FF000000"),
        # )

        ws.merge_cells("A1:F2")
        top_left_cell = ws["A1"]
        top_left_cell.value = "Karta Pracy"
        # top_left_cell.border = border

        wb.save(filename=wk_name)
