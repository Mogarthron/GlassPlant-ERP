import datetime
from datetime import timedelta, datetime, date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, PatternFill
import calendar
import pandas as pd
from pandas import DataFrame, Series
from DBConnection import DBConnection
import os.path


class ScheduleModel():

    def __init__(self):

        self.MonthNames = ['Styczeń', 'Luty', 'Marzec', 'Kwiecień', 'Maj', 'Czerwiec',
                           'Lipiec', 'Sierpień', 'Wrzesień', 'Październik', 'Listopad', 'Grudzień']

        self.weekDay = ['poniedziałek', 'wtorek', 'środa',
                        'czwartek', 'piątek', 'sobota', 'niedziela']

        self.r = ['Ranek', 'R']
        self.p = ['Popołudnie', 'P']
        self.n = ['Nocka', 'N']
        self.w = ['Wolne', 'W']
        self.r7 = ['Ranek od 7:00', '7']
        self.r8 = ['Ranek od 8:00', '8']
        self.wn = ['Wolna niedziela', 'wn']
        self.dt = ['Dzień wolny wynikający z grafiku', 'dt']
        self.h = ['Święta i dni wolne od pracy', 'h']
        self.l4 = ['Zwolnienie lekarskie', 'l4']
        self.kw = ['Kwarantanna', 'kw']
        self.uw = ['Urlop wypoczynkowy', 'uw']
        self.uo = ['Urlop okolicznościowy', 'uo']
        self.ub = ['Urlop bezpłatny', 'ub']
        self.nn = ['Nieobecność nieusprawiedliwiona', 'nn']
        self.nu = ['Nieobecność usprawiedliwiona', 'nu']

        # list of schedule components
        self.sc = [
            self.r,
            self.p,
            self.n,
            self.w

        ]

    def SchedulePatern(self, patern):
        '''patern is string like RRRRWPPPPWNNNNWW '''

        _schedulePatern = list()

        for i in patern:
            if(i == 'R'):
                _schedulePatern.append(self.r)
            elif (i == 'P'):
                _schedulePatern.append(self.p)
            elif (i == 'N'):
                _schedulePatern.append(self.n)
            elif (i == 'W'):
                _schedulePatern.append(self.w)
            elif (i == '7'):
                _schedulePatern.append(self.r7)
            elif (i == '8'):
                _schedulePatern.append(self.r7)

        return _schedulePatern

    def WeekDay(self, data):
        r = list()
        r.append(data)
        r.append(self.weekDay[data.weekday()])
        return r

    def MonthDay(self, year, month):

        data = date(year, month, 1)

        monthDay = list()

        if (data.month == 12):
            while (data.year < year + 1):
                monthDay.append(self.WeekDay(data))
                data = data + timedelta(days=1)
        else:
            while (data.month < month + 1):
                monthDay.append(self.WeekDay(data))
                data = data + timedelta(days=1)

        return monthDay

    def SetShiftsDiuringMonth(self, year, month, schedulePatern, startShift):
        '''
        year: int type

        month: int type

        schedulePatern: string like RRRRWPPPPWNNNNWW

        startShift: position of first day of shedule from pattern, int type
        '''
        shift = list()

        for __d in self.MonthDay(year, month):

            shift.append(schedulePatern[startShift])

            startShift = startShift + 1
            if (len(schedulePatern) - 1 < startShift):
                startShift = 0

        # return [x[1] for x in shift]
        return shift


class WorkSchedule():

    def __init__(self, department, year, month):
        '''year and month mast by int type'''

        self.__sm = ScheduleModel()

        self.Department = department
        self.year = year
        self.month = month

        # self.MonthNames = self.__sm.MonthNames
        # self.weekDay = self.__sm.weekDay

        self.__departmentName = 'Wydział: '

    def __SetListOfEmployes(self):

        parms = [self.year, self.month]
        query = 'exec [GPERP].[dbo].[spPokazHarmonogram_HarmonogramPracyPracownika] ? , ?'
        dbcon = DBConnection()
        listOfEmployes = dbcon.ShowQuerry(query, parms)

        return listOfEmployes

    def SetShiftsDiuringMonth(self, schedulePatern, startShift):

        shift = list()

        for __d in self.__sm.MonthDay(self.year, self.month):

            shift.append(schedulePatern[startShift])

            startShift = startShift + 1
            if (len(schedulePatern) - 1 < startShift):
                startShift = 0

        return [x[1] for x in shift]

    def CreateSchedule(self):

        sh = DataFrame()

        sh['Dzień mc'] = [x[0]
                          for x in self.__sm.MonthDay(self.year, self.month)]
        sh['Dzień tyg'] = [x[1]
                           for x in self.__sm.MonthDay(self.year, self.month)]

        __e = self.__SetListOfEmployes()

        for i in __e.index:
            sh[__e['Pracownik'][i]] = self.SetShiftsDiuringMonth(
                self.__sm.SchedulePatern(__e['przebiegZmian'][i]), __e['poczatkowaZmiana'][i])

        sh.to_excel('Harmonogram_Template.xlsx', index=False)
        print(sh.to_string(index=False))

    def ReadShedule(self):

        df = pd.read_excel('Harmonogram_Template.xlsx', engine='openpyxl')

        # sheduleName = self.Department + '_' + \
        #     str(self.year) + '_' + str(self.month) + '.xlsx'

        return df
        # print(df)
        # df.to_excel(sheduleName)

        # if os.path.isfile('Harmonogram_Template.xlsx'):
        #     os.remove('Harmonogram_Template.xlsx')


class WorkCard():

    def __init__(self, department, year, month):

        self.year = year
        self.month = month

        self.sm = ScheduleModel()
        self.__dep = department
        self.__departmentName = 'Wydział: ' + department

        self.__db = DBConnection()

        self.__ws = list()

        self.__querys = ['exec [GPERP].[dbo].[spDaneDoHarmonogramu] ?, ?, ?',
                         'exec [GPERP].[dbo].[spDniWolneOdPracyWynikajaceZHarmonogramu] ?, ?, ?',
                         'exec [GPERP].[dbo].[spNieobecnosciPracownika] ?, ?, ?',
                         'exec [GPERP].[dbo].[spDniWolneOdPracyISwieta] ?, ?']

        self.__holidays = self.__db.ShowQuerry(self.__querys[3], [year, month])

        self.__holidays = self.__holidays['data'].to_list()

        # convert output from querry to date type
        for i in range(len(self.__holidays)):
            self.__holidays[i] = datetime.strptime(
                self.__holidays[i], '%Y-%m-%d').date()

        # Table of all employs from department
        self.__emp = self.__db.ShowQuerry(
            self.__querys[0], [self.__dep, self.year, self.month])

    def ShowEmployData(self, empName=''):
        '''
        For empty empName return dataframe __emp which contain employs  info about salary and shedule for specific month.
        If empName contain string or int function will return secific row from dataframe __emp
        '''

        if (empName == ''):
            return self.__emp

        if (empName != '' and type(empName) == str):
            dataForWorkshedule = self.__emp[self.__emp['pracownik'] == empName]

            employ = [dataForWorkshedule['pracownik'].values[0],
                      dataForWorkshedule['stawka'].values[0],
                      dataForWorkshedule['nazwaStanowiska'].values[0],
                      dataForWorkshedule['poczatkowaZmiana'].values[0],
                      dataForWorkshedule['przebiegZmian'].values[0]]

            return employ

        if (empName != '' and type(empName) == int):
            dataForWorkshedule = self.__emp.loc[empName]
            employ = [dataForWorkshedule['pracownik'],
                      dataForWorkshedule['stawka'],
                      dataForWorkshedule['nazwaStanowiska'],
                      dataForWorkshedule['poczatkowaZmiana'],
                      dataForWorkshedule['przebiegZmian']]

            return employ

    def SetEmploySheduleForWorkCard(self, empName):
        '''
        Function return workshedule for one employ with all unpresence during month.

        empName: Surname and First name of employ or number of row in __emp dataframe
        '''
        self.__ws = list()

        # choose one specific employ from emp table
        employ = self.ShowEmployData(empName)
        ############################################################

        NoneWorkingDays = self.__db.ShowQuerry(
            self.__querys[1], [self.year, self.month, employ[0]])

        DaysOff = self.__db.ShowQuerry(
            self.__querys[2], [self.year, self.month, employ[0]])

        ### Set basic work schedule ###########################
        sp = self.sm.SchedulePatern(employ[4])
        self.__ws = self.sm.SetShiftsDiuringMonth(
            self.year, self.month, sp, employ[3])

        ######################################################

        if (employ[2] != 'Topiarz'):

            for i in range(len(self.__holidays)):
                self.__ws[self.__holidays[i].day - 1] = self.sm.h

        t = NoneWorkingDays['skrot'].to_list()  # type of none working day

        d = NoneWorkingDays['data'].to_list()  # date of none working day

        for i in range(len(t)):

            if (t[i] == self.sm.h[1]):
                self.__ws[datetime.strptime(
                    d[i], '%Y-%m-%d').day - 1] = self.sm.h
            elif (t[i] == self.sm.dt[1]):
                self.__ws[datetime.strptime(
                    d[i], '%Y-%m-%d').day - 1] = self.sm.dt

        dotype = DaysOff['skrot'].to_list()  # type of days of like l4, uw, uo
        dofrom = DaysOff['nieobecnoscOd'].to_list()  # daysof from date
        doto = DaysOff['nieobecniscDo'].to_list()  # daysof to date
        # how meny days unpresens taken
        doduration = DaysOff['czasNieprzeracowany'].to_list()

        for i in range(len(dotype)):

            # dm: date of dayof converted to date time type
            dm = datetime.strptime(dofrom[i], '%Y-%m-%d')

            if (dotype[i] == self.sm.uw[1]):

                self.__DayOffInWorkSchedule(
                    self.sm.uw, doduration[i], dm, doto[i], dofrom[i])

            if (dotype[i] == self.sm.kw[1]):

                self.__OverridableDaysOff(
                    self.sm.kw, doduration[i], dm, doto[i], dofrom[i])

            if (dotype[i] == self.sm.l4[1]):

                self.__OverridableDaysOff(
                    self.sm.l4, doduration[i], dm, doto[i], dofrom[i])

        return self.__ws

    def __DayOffInWorkSchedule(self, sm_element, doduration, dm, doto, dofrom):
        '''
        void change work shedule positions to upsence elements

        sm_element: element from SheduleModel class 

        doduration: list of int's describe 

        dm:

        doto:

        dofrom:

        '''

        b = 0

        if (doduration == 8 and self.__ws[dm.day - 1] != self.sm.w and self.__ws[dm.day - 1] != self.sm.dt):
            self.__ws[dm.day - 1] = sm_element

        else:
            while (dm < datetime.strptime(doto, '%Y-%m-%d')):

                dm = datetime.strptime(dofrom, '%Y-%m-%d') + timedelta(b)

                if (self.__ws[dm.day - 1] != self.sm.w and self.__ws[dm.day - 1] != self.sm.dt and self.__ws[dm.day - 1] != self.sm.h):
                    self.__ws[dm.day - 1] = sm_element

                b = b + 1

    def __OverridableDaysOff(self, sm_element, doduration, dm, doto, dofrom):
        '''

        '''

        b = 0

        if (doduration == 8 and self.__ws[dm.day - 1] == self.sm.w and self.__ws[dm.day - 1] == self.sm.dt):
            self.__ws[dm.day - 1] = sm_element

        else:
            while (dm < datetime.strptime(doto, '%Y-%m-%d')):

                dm = datetime.strptime(
                    dofrom, '%Y-%m-%d') + timedelta(b)

                self.__ws[dm.day - 1] = sm_element

                b = b + 1

    def PrintWorkCardToExcell(self, Employ, workSchedule, Holidays):

        wb = load_workbook(filename='./Resources/Templates/Karta_pracy.xlsx')
        sheet = wb.active

        sheet['A3'] = self.__departmentName
        sheet['G1'] = self.sm.MonthNames[self.month-1]
        sheet['G2'] = self.year
        sheet['M1'] = Employ[0]  # Surname and name of employ
        sheet['AC1'] = Employ[1]  # Salary
        sheet['AI1'] = Employ[2]  # Position of employment

        colNum = 4

        workTime = 0
        nightWork = 0
        vacationLeave = 0  # Urlop wypoczynkowy
        sickLeave = 0  # zwolnienie lekarskie

        if (Employ[2] == 'Topiarz'):
            sheet['A8'] = 'Topienie szkła'
            sheet['A9'] = 'Wylewanie szkła opalowego'
        else:
            sheet['A8'] = 'Czas przepracowany'

        for d in range(calendar.monthrange(self.year, self.month)[1]):

            # style cel in date row if day is a holiday
            for i in Holidays:
                if (date(self.year, self.month, d+1) == i):
                    CellFontColor = Font(
                        color='00FFFFFF', name='Calibri', size=11)
                    CellFillColor = PatternFill("solid", fgColor='00808080')
                    c = sheet.cell(row=5, column=colNum, value=d + 1)
                    c.font = CellFontColor
                    c.fill = CellFillColor
                else:
                    sheet.cell(row=5, column=colNum, value=d + 1)

            # style cell in date row if day is Sunday
            if (date(self.year, self.month, d+1).weekday() == 6):
                CellFontColor = Font(color='00FFFFFF', name='Calibri', size=11)
                CellFillColor = PatternFill("solid", fgColor='00808080')
                c = sheet.cell(row=5, column=colNum, value=d + 1)
                c.font = CellFontColor
                c.fill = CellFillColor
            else:
                sheet.cell(row=5, column=colNum, value=d + 1)

            # Ranki / morning shift
            if (workSchedule[d] == self.sm.r):
                self.__ShiftRow(sheet, 6, 14, colNum)
                workTime = workTime + 8
            # Popołudnia / afternoon shift
            elif (workSchedule[d] == self.sm.p):
                self.__ShiftRow(sheet, 14, 22, colNum)
                workTime = workTime + 8
            # Nocki / night shift
            elif (workSchedule[d] == self.sm.n):
                self.__ShiftRow(sheet, 22, 6, colNum)
                sheet.cell(row=12, column=colNum, value=8)
                workTime = workTime + 8
                nightWork = nightWork + 8
            # Ranki Od 7:00
            elif (workSchedule[d] == self.sm.r7):
                self.__ShiftRow(sheet, 7, 15, colNum)
                workTime = workTime + 8
            # Ranki Od 8:00
            elif (workSchedule[d] == self.sm.r8):
                self.__ShiftRow(sheet, 8, 16, colNum)
                workTime = workTime + 8
            # Wolne i wolne wynikające z grafiku
            elif (workSchedule[d] == self.sm.w or workSchedule[d] == self.sm.dt):
                sheet.cell(row=6, column=colNum, value='')
            # Urlop wypoczynkowy
            elif (workSchedule[d] == self.sm.uw):
                sheet.cell(row=13, column=colNum, value=8)
                sheet.cell(row=23, column=colNum, value=8)
                vacationLeave = vacationLeave + 8
            # Urlop okolicznościowy
            elif (workSchedule[d] == self.sm.uo):
                sheet.cell(row=17, column=colNum, value=8)
                sheet.cell(row=23, column=colNum, value=8)
            # Zwolnienie Lekarskie
            elif (workSchedule[d] == self.sm.l4):
                sheet.cell(row=18, column=colNum, value=8)
                sheet.cell(row=23, column=colNum, value=8)
                sickLeave = sickLeave + 8

            colNum = colNum + 1

        #######################################################################

        sheet['AI8'] = workTime

        if (nightWork != 0):
            sheet['AI12'] = nightWork

        if (vacationLeave != 0):
            sheet['AI13'] = vacationLeave

        if (sickLeave != 0):
            sheet['AI18'] = sickLeave

        sheet['AI23'] = workTime + vacationLeave + sickLeave

        #########################################################################

        wk_name = './Resources/Output/' + Employ[0] + '.xlsx'

        if os.path.isfile(wk_name):
            os.remove(wk_name)

        print('Printing...')
        wb.save(wk_name)
        print('Work card printed')

    def __ShiftRow(self, sheet, shiftfrom, shiftto, colNum):

        sheet.cell(row=6, column=colNum, value=shiftfrom)
        sheet.cell(row=7, column=colNum, value=shiftto)
        sheet.cell(row=8, column=colNum, value=8)
        sheet.cell(row=23, column=colNum, value=8)
##########################################################################################################
##########################################################################################################
